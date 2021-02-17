# # This is Base class for reusability for facebook url
# from selenium import webdriver
# #
# #
# class Base:
#     def launch_browser(self):
#         self.driver = webdriver.Chrome(executable_path=r"C:"
#                                                   r"\Users\Admin\PycharmProjects\ferozproject\Chrome\chromedriver.exe")
#         self.driver.maximize_window()
#         return self.driver
#
#     def load_url(self, url):  # pass the current self instance of object an argument url
#         x=self.driver.get(url)
#         return x
#
#     def type(self, element, data):
#         element.send_keys(data)
#
#     def btn_click(self,e):
#         e.click()
#
#     # def scrolldown_page(self,element):
#     #     element.
#     def btn_create(self,e):
#         e.click()
#
#     # # btn_create.click()
# ************************************************************
#base class for the url adcatinehotel url
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
import openpyxl
class Base:
    def launch_browser(self):
        self.driver=webdriver.Chrome(executable_path=r"C:\Users\Admin\PycharmProjects\ferozproject\Chrome\chromedriver.exe")
        self.driver.maximize_window()
        return self.driver
    def load_url(self,url):
        self.driver.get(url)

    def insert(self,element,data):
        element.send_keys(data)

    def btn_login(self,e):
        e.click()

    def current_url(self):
        y = self.driver.current_url
        return y

    def type(self, element, data):
        element.send_keys(data)

    def btn_click(self, ele):
        ele.click()

    def select_by_index(self, element, index):
        s = Select(element)
        s.select_by_index(index)

    def select_by_value(self,element,value):
        s=Select(element)
        s.select_by_value(value)


    def select_by_visible_text(self,element,visibletext):
        s=Select(element)
        s.select_by_visible_text(visibletext)
        s.deselect_by_visible_text(visibletext)




    #  mouse_overAction methods


    def move_to_element(self, e):
        acc = ActionChains(self.driver)
        acc.move_to_element(e).perform()

    def drag_and_drop(self,a,b):
        acc = ActionChains(self.driver)
        acc.drag_and_drop(a,b).perform()

    # keyborad Actions
    def context_click(self,e):
        acc=ActionChains(self.driver)
        acc.context_click(e).perform()

    def double_click(self,e):
        acc=ActionChains(self.driver)
        acc.double_click(e).perform()

    def btn_click(self, e):
        e.click()

    # # btn_create.click()
    def get_data_from_excel(self,excel_loc,row_index,cell_index):
        w= openpyxl.loadworksheet(excel_loc)
        sheet=w.active
        cell=sheet.cell(row_index,cell_index)
        v=cell.value
        return v


    def quit_browser(self):
        self.driver.quit()

    def page_tittle(self):
        t = self.driver.title
        return t

################################################################
class Main:
    def launch_browser(self):
        self.driver=webdriver.Chrome(executable_path=r"C:\Users\Admin\PycharmProjects\ferozproject\Chrome\chromedriver.exe")
        self.driver.maximize_window()
        return self.driver
    def load_url(self,url):
        self.driver.get(url)
    def insert(self,element,data):
        element.send_keys(data)
    def btn_login(self,element):
        element.click()

    def select_by_index(self,element,index):
        s=Select(element)
        s.select_by_index(index)
    def select_by_value(self,element,value):
        s=Select(element)
        s.select_by_value(value)
    def select_by_visibletext(self,element,visibletext):
        s=Select(element)
        s.select_by_visible_text(visibletext)
    def move_to_element(self,element):
        acc=ActionChains(self.driver)
        acc.move_to_element(element).perform()
    def drag_and_drop(self,a,b):
        acc=ActionChains(self.driver)
        acc.drag_and_drop(a,b).perform()
    def context_click(self,element):
        acc=ActionChains(self.driver)
        acc.context_click(element).perform()
    def double_clcick(self,element):
        acc=ActionChains(self.driver)
        acc.double_click(element).perform()
    def get_data_from_excel(self,excel_loc,row_index,cell_index):
        w=openpyxl.load_workbook(excel_loc)
        sheet=w.active
        cell=sheet.cell(row_index,cell_index)
        v= cell.value
        return v
