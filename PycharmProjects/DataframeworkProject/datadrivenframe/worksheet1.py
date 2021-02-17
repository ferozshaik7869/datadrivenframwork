# Data Driven framework
# Excel:
# xlxs
# xls
# send_keys("")
# 1)install package openpyxl
# 2)import openpyxl
# 3)load workbook
import openpyxl

#mention location of the excel
excel_loc=r"C:\Users\Admin\Desktop\feroz.xlsx"
#load workbook
w=openpyxl.load_workbook(excel_loc)
# to get the sheet from workbook
sheet=w.active
# to get cell from the sheet
cell=sheet.cell(1,1)
#to get value from the cell
v=cell.value
print(v)

import openpyxl
excel_loc=r"C:\Users\Admin\Desktop\hello.xlsx"
w=openpyxl.load_workbook(excel_loc)
sheet=w.active
cell=sheet.cell(1,1)
v=cell.value
print(v)

import openpyxl
location=r"C:\Users\Admin\Desktop\hello.xlsx"
w=openpyxl.load_workbook(location)
sheet=w.active
cell=sheet.cell(2,2)
v=cell.value
print(v)

import openpyxl
loc=r"C:\Users\Admin\Desktop\hello.xlsx"
w=openpyxl.load_workbook(loc)
sheet=w.active
cell=sheet.cell(1,2)
v=cell.value
print(v)

import openpyxl
locat=r"C:\Users\Admin\Desktop\hello.xlsx"
w=openpyxl.load_workbook(locat)
sheet=w.active
cell=sheet.cell(2,1)
v=cell.value
print(v)

import openpyxl
loca=r"C:\Users\Admin\Desktop\hello.xlsx"
w=openpyxl.load_workbook(loca)
sheet=w.active
cell=sheet.cell(1,3)
v=cell.value
print(v)

import openpyxl
locate=r"C:\Users\Admin\Desktop\hello.xlsx"
w=openpyxl.load_workbook(locate)
sheet=w.active
cell=sheet.cell(2,2)
v=cell.value
print(v)

# # **********************************************
# # # practice of above program
import openpyxl
#mention the location of the excel
excel_loc1=r"C:\Users\Admin\Desktop\feroz.xlsx"
#load work book
w1=openpyxl.load_workbook(excel_loc1)
# to get the sheet from workbook
sheet=w1.active
#to get cell from the sheet

cell=sheet.cell(2,2)
#to get value from the cell
value=cell.value
print(value)
#
r=sheet.max_row
print(r)

c=sheet.max_column
print(c)
# *********************************************
# for r in range(1, r+1):
#     for c in range(1,c+1):
#         c1=sheet.cell(r,c)
#         data=c1.value
#         print(data)
# # *************************************************
# import openpyxl
# excel_loc=r"C:\Users\Admin\Desktop\feroz.xlsx"
# w=openpyxl.load_workbook(excel_loc)
# sheet=w.active
# cell=sheet.cell(2,2)
# value=cell.value
# print(value)
#
# for r in range(1, r+1):
#     for c in range(1,c+1):
#         c1=sheet.cell(r,c)
#         data=c1.value
#         print(data)

# *******************************************************
# create a sheetin workbook is as follows
import openpyxl
from openpyxl.workbook.workbook import save_workbook


excel_loc="shaik.xlsx"
#create a workbook
w = openpyxl.Workbook()
#create a sheet
sheet=w.create_sheet("hello")
# adding value or appending value in the sheet
sheet.cell(1,1).value="java"

# save the file which is created
w.save(excel_loc)
# **********************************************
import openpyxl
excel_loc=r"C:\Users\Admin\Desktop\kursheed.xlsx"
w=openpyxl.load_workbook(excel_loc)
sheet=w.active
cell=sheet.cell(1,1)
r=sheet.max_row
c=sheet.max_column
sheet=w.active

value=cell.value
print(value)
for r in range(1,r+1):
    for c in range(1,c+1):
        c1=sheet.cell(r,c)
        data=c1.value
        print(data)

